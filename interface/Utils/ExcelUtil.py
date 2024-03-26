import openpyxl

from interface.Conf.settings import TEST_DATA_DIR


def get_row_values(file_name, sheet_name,row):
    """
    获取Excel标题行，就是获取入参名字
    """
    file_path = TEST_DATA_DIR + f'/{file_name}'
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    columns = sheet.max_column
    data = []
    for c in range(1, columns+1):
        data.append(sheet.cell(row, c).value)
    return data

def get_test_data(file_name, sheet_name):
    file_path = TEST_DATA_DIR + f'/{file_name}'
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    row_data = []
    column_data = []
    rows = sheet.max_row
    columns = sheet.max_column
    for i in range(2, rows+1):
        row_data = []
        column_data.append(row_data)
        for j in range(1, columns+1):
            row_data.append(sheet.cell(row=i, column=j).value)
    return column_data


if __name__ == '__main__':
    title = get_row_values('data.xlsx', 'Sheet1', 1)
    print(title)
    data = get_test_data('data.xlsx', 'Sheet1')
    print(data)

