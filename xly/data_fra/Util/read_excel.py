# -*- coding: utf-8 -*- 
# @Time : 2023-12-17 14:50 
# @Author : kunp
# @File : read_excel.py 
# @Software: PyCharm
from openpyxl import load_workbook

from data_fra.Settings.Config import testDataPath


class ExcelParse:

    def __init__(self):
        self.workbook = None
        self.sheet = None


    def load_workbook(self, filename):
        """
        加载数据文件
        :param filename:
        :return:
        """
        try:
            self.workbook = load_workbook(filename)
        except Exception as e:
            raise e

    def get_sheet_by_name(self, sheetname):
        """
        获取具体的sheet
        :param sheetname:
        :return:
        """
        try:
            self.sheet = self.workbook[sheetname]
        except Exception as e:
            raise e

    def get_cell_value(self, row, column):
        """
        获取某个单元格的值
        :param row:
        :param column:
        :return:
        """
        try:
            cell_value = self.sheet.cell(row=row, column=column).value
            return cell_value
        except Exception as e:
            raise e

    def get_row_value(self, row):
        """
        获取某一行的值
        :param row:
        :return:
        """
        # 获取列数
        try:
            columns = self.sheet.max_column
            row_data = []
            for i in range(1, columns+1):
                cell_value = self.sheet.cell(row=row, column=i).value
                row_data.append(cell_value)
            return row_data
        except Exception as e:
            raise e

    def get_rows_nums(self):
        """
        获取行数
        :return:
        """
        return self.sheet.max_row

    def get_columns_nums(self):
        """
        获取列数
        :return:
        """
        return  self.sheet.max_column


if __name__ == '__main__':
    excel_par = ExcelParse()
    excel_par.load_workbook(testDataPath)
    excel_par.get_sheet_by_name('login')
    print(excel_par.get_cell_value(2, 1))
    print(excel_par.get_row_value(2))
    print(excel_par.get_rows_nums())
    print(excel_par.get_columns_nums())
