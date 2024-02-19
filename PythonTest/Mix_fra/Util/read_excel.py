import os

from openpyxl import load_workbook
#openpyxl 从1开始计数
from PythonTest.Mix_fra.settings.config import testDataPath
# 改造读取文件

class ExcelParse:

    def __init__(self):
        self.workbook = None   # 是哪一个文件
        #self.sheet = None  # Excel中sheet
        # 由于涉及到Data 和 keyword两份sheet 所以不初始化sheet

    def load_workbook(self, filename):
        """
        加载数据文件
        :param filename:
        :return:
        """
        try:
            self.workbook = load_workbook(filename)
        except Exception as e:
            raise e

    def get_sheet_by_name(self, sheetname):
        """
        获取具体的sheet
        :param sheetname:
        :return:
        """
        try:
            return self.workbook[sheetname]
        except Exception as e:
            raise e

    def get_cell_value(self, row, column, sheet):
        """
        获取某一单元格的值
        :param row: 行数
        :param column: 列数
        :return:
        """
        try:
            cell_value = sheet.cell(row=row, column=column).value
            return cell_value
        except Exception as e:
            raise e

    def get_row_value(self, row, sheet):
        """
        获取某一行的值
        :param row:
        :return:
        """
        try:
            columns = sheet.max_column
            row_data = []
            for i in range(1, columns + 1):
                cell_value = sheet.cell(row=row, column=i).value
                row_data.append(cell_value)
            return row_data
        except Exception as e:
            raise e

    def get_rows_nums(self, sheet):
        """
        获取行数
        :return:
        """
        return sheet.max_row

    def get_columns_nums(self, sheet):
        """
        获取列数
        :return:
        """
        return sheet.max_column

    def write_cell(self, row, column, value, file_name, sheet):
        """
        写入Excel单元格 中执行结果
        :return:
        """
        try:
            sheet.cell(row=row, column=column, value=value)
            self.workbook.save(file_name)
        except Exception as e:
            raise e


# if __name__ == "__main__":
#     excel_par = ExcelParse()
#     excel_par.load_workbook(os.path.join(testDataPath, "163mail.xlsx"))
#     excel_par.get_sheet_by_name("index")
#     excel_par.write_cell(2, 6, "pass", os.path.join(testDataPath, "163mail.xlsx"), excel_par.sheet)
#     excel_par.load_workbook(os.path.join(testDataPath, "163mail.xlsx"))
#     excel_par.get_sheet_by_name('login')
#     print(excel_par.get_cell_value(2, 1))
#     print(excel_par.get_row_value(2))
#     print(excel_par.get_rows_nums())
#     print(excel_par.get_columns_nums())
